"""Attachment parsers for the Local Knowledge Store.

Parses attachment files into extract JSON sidecars.  Every attachment is
captured — silently dropping nothing.  Unparseable attachments get a
metadata-only extract with needsReview=true.
"""

from __future__ import annotations

import csv
import json
import os
import tempfile
import uuid
from datetime import datetime, timezone
from typing import Any

from .hashing import sha256_text, sha256_bytes, make_extract_key


# ── Temporary file helpers ─────────────────────────────────────────────

TEMP_PARSING_DIR = "temp/parsing"


def _ensure_temp_dir(store_root: str) -> str:
    p = os.path.join(store_root, TEMP_PARSING_DIR)
    os.makedirs(p, exist_ok=True)
    return p


def _save_temp_attachment(store_root: str, com_attachment: Any) -> tuple[str, str, str]:
    """Save a COM attachment to temp, return (temp_path, original_name, extension)."""
    temp_dir = _ensure_temp_dir(store_root)
    orig_name = str(com_attachment.FileName) if hasattr(com_attachment, "FileName") else "unknown"
    ext = os.path.splitext(orig_name)[1].lower()
    temp_name = f"{uuid.uuid4().hex}{ext}"
    temp_path = os.path.join(temp_dir, temp_name)
    com_attachment.SaveAsFile(temp_path)
    return temp_path, orig_name, ext


def _delete_temp_file(path: str) -> bool:
    """Delete a temp file, return True if successful."""
    try:
        if os.path.isfile(path):
            os.remove(path)
            return not os.path.isfile(path)
    except Exception:
        pass
    return False


# ── Parser dispatch ───────────────────────────────────────────────────


def _try_openpyxl(path: str) -> dict[str, Any] | None:
    """Try openpyxl for .xlsx/.xlsm files."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                sheets.append({
                    "name": name,
                    "rowCount": len(rows),
                    "cells": [[str(c) if c is not None else "" for c in row] for row in rows],
                })
        wb.close()
        return {
            "status": "parsed",
            "parser": "openpyxl",
            "metadata": {"sheetCount": len(sheets), "sheets": sheets},
        }
    except Exception:
        return None


def _try_python_csv(path: str) -> dict[str, Any] | None:
    """Try csv module for .csv files."""
    try:
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = [[c.strip() for c in row] for row in reader]
        text = "\n".join("\t".join(row) for row in rows)
        return {
            "status": "parsed",
            "parser": "csv",
            "text": text,
            "metadata": {"rowCount": len(rows)},
        }
    except Exception:
        return None


def _try_text(path: str) -> dict[str, Any] | None:
    """Direct text extraction for .txt, .log, .xml, .json, etc."""
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            text = f.read()
        ext = os.path.splitext(path)[1].lower()
        return {
            "status": "parsed",
            "parser": f"text_{ext[1:]}" if ext else "text",
            "text": text,
        }
    except Exception:
        return None


def _try_docx(path: str) -> dict[str, Any] | None:
    """Try python-docx for .docx files."""
    try:
        from docx import Document
        doc = Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
        return {
            "status": "parsed",
            "parser": "python-docx",
            "text": text,
            "metadata": {"paragraphCount": len(paragraphs)},
        }
    except ImportError:
        return {
            "status": "metadata_only",
            "parser": "python-docx",
            "failureReason": "parser_dependency_unavailable",
            "needsReview": True,
        }
    except Exception:
        return None


def _try_pdf(path: str) -> dict[str, Any] | None:
    """Try PyMuPDF for .pdf files."""
    try:
        import fitz
        doc = fitz.open(path)
        pages = []
        for page_num in range(len(doc)):
            pages.append(doc[page_num].get_text())
        text = "\n".join(pages)
        doc.close()
        return {
            "status": "parsed",
            "parser": "PyMuPDF",
            "text": text,
            "metadata": {"pageCount": len(pages)},
        }
    except ImportError:
        return {
            "status": "metadata_only",
            "parser": "PyMuPDF",
            "failureReason": "parser_dependency_unavailable",
            "needsReview": True,
        }
    except Exception:
        return None


# ── Main parser ────────────────────────────────────────────────────────


def parse_attachment(
    store_root: str,
    com_attachment: Any,
    parent_record_key: str,
    export_run_id: str,
) -> dict[str, Any]:
    """Parse a COM attachment and return an extract dict.

    The temp file is saved, parsed, then deleted.  Never retains raw files.
    """
    temp_path, orig_name, ext = _save_temp_attachment(store_root, com_attachment)
    orig_name_hash = sha256_text(orig_name)
    size = os.path.getsize(temp_path) if os.path.isfile(temp_path) else 0
    content_hash = sha256_bytes(open(temp_path, "rb").read()) if os.path.isfile(temp_path) else ""

    result: dict[str, Any] | None = None

    # Dispatch by extension
    if ext in (".xlsx", ".xlsm"):
        result = _try_openpyxl(temp_path)
    elif ext == ".csv":
        result = _try_python_csv(temp_path)
    elif ext in (".txt", ".log", ".xml", ".json", ".md", ".yaml", ".yml", ".ini", ".cfg", ".py", ".js", ".html", ".css"):
        result = _try_text(temp_path)
    elif ext == ".docx":
        result = _try_docx(temp_path)
    elif ext == ".pdf":
        result = _try_pdf(temp_path)
    else:
        # Unknown/binary — metadata-only
        result = {
            "status": "metadata_only",
            "parser": "none",
            "failureReason": "unsupported_type",
            "needsReview": True,
        }

    # If result is None, parser ran but failed
    if result is None:
        result = {
            "status": "failed",
            "parser": "none",
            "failureReason": "parser_failed",
            "needsReview": True,
        }

    # Build extract data
    parse_status = result.get("status", "failed")
    parser_name = result.get("parser", "none")
    failure_reason = result.get("failureReason", "")
    needs_review = result.get("needsReview", False)
    text = result.get("text", "")
    text_hash = sha256_text(text) if text else ""
    parse_metadata = result.get("metadata", {})

    extract_key = make_extract_key(parent_record_key, orig_name_hash)
    now_iso = datetime.now(timezone.utc).isoformat()

    # Write extract sidecar
    extract = {
        "_schema": "export.knowledgeExtract.v1",
        "recordType": "outlookAttachmentExtract",
        "parentRecordKey": parent_record_key,
        "exportRunId": export_run_id,
        "extractKey": extract_key,
        "source": {
            "originalName": orig_name,
            "originalNameHash": orig_name_hash,
            "extension": ext,
            "sizeBytes": size,
            "contentHash": content_hash,
        },
        "parse": {
            "status": parse_status,
            "parser": parser_name,
            "parsedAt": now_iso,
            "failureReason": failure_reason,
            "needsReview": needs_review,
        },
        "content": {
            "text": text,
            "textHash": text_hash,
            "tables": [],
            "sheets": parse_metadata.get("sheets", []),
            "metadata": parse_metadata,
        },
        "retrieval": {"chunkIds": []},
        "vault": {"notePaths": [], "canvasPaths": []},
        "audit": {
            "rawSourceRetained": False,
            "tempFileDeleted": True,
            "parseWarnings": [],
        },
    }

    # Write to disk
    year, month = now_iso[:4], now_iso[5:7]
    extract_dir = os.path.join(store_root, "extracts", year, month)
    os.makedirs(extract_dir, exist_ok=True)
    extract_path = os.path.join(
        extract_dir,
        f"record_{parent_record_key}_extract_1_{extract_key}.json",
    )
    with open(extract_path, "w", encoding="utf-8") as f:
        json.dump(extract, f, indent=2, ensure_ascii=False)

    # Delete temp file
    deleted = _delete_temp_file(temp_path)
    extract["audit"]["tempFileDeleted"] = deleted
    if not deleted:
        extract["audit"]["parseWarnings"].append("temp_file_not_deleted")

    return {
        "extract": extract,
        "extract_key": extract_key,
        "extract_path": extract_path,
        "parse_status": parse_status,
        "needs_review": needs_review,
        "temp_deleted": deleted,
    }
