"""Tests for parsers.py — attachment parsing."""

import json
import os
import tempfile

from export_engine.parsers import (
    parse_attachment,
    _try_text,
    _try_python_csv,
    _try_openpyxl,
)


class TestTextParser:
    def test_txt_extraction(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "test.txt")
            with open(path, "w") as f:
                f.write("Hello, world!\nLine 2")
            result = _try_text(path)
            assert result is not None
            assert result["status"] == "parsed"
            assert "Hello, world!" in result["text"]

    def test_json_extraction(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "test.json")
            with open(path, "w") as f:
                json.dump({"key": "value"}, f)
            result = _try_text(path)
            assert result is not None
            assert result["status"] == "parsed"

    def test_xml_extraction(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "test.xml")
            with open(path, "w") as f:
                f.write("<root><item>val</item></root>")
            result = _try_text(path)
            assert result is not None
            assert result["status"] == "parsed"


class TestCsvParser:
    def test_csv_extraction(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "test.csv")
            with open(path, "w", newline="") as f:
                f.write("name,value\nAlice,100\nBob,200")
            result = _try_python_csv(path)
            assert result is not None
            assert result["status"] == "parsed"
            assert "Alice" in result["text"]
            assert "Bob" in result["text"]


class TestOpenpyxlParser:
    def test_xlsx_if_available(self) -> None:
        try:
            import openpyxl
        except ImportError:
            return  # Skip if not available
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "test.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Hello"
            ws["B1"] = 42
            wb.save(path)
            result = _try_openpyxl(path)
            assert result is not None
            assert result["status"] == "parsed"
            assert result["metadata"]["sheetCount"] >= 1
            sheets = result["metadata"]["sheets"]
            assert sheets[0]["cells"][0][0] == "Hello"


class TestParseAttachmentFlow:
    """Full attachment parse-save-temp-delete flow."""

    def _make_dummy_com_attachment(self, store_root: str) -> object:
        """Create a minimal COM-like attachment mock."""
        import types

        class MockAttachment:
            def __init__(self):
                self.FileName = "test.txt"
            def SaveAsFile(self, path):
                with open(path, "w") as f:
                    f.write("Hello attachment!")

        return MockAttachment()

    def test_parse_and_delete_temp(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store_root = os.path.join(tmp, "KnowledgeStore")
            os.makedirs(os.path.join(store_root, "temp", "parsing"), exist_ok=True)
            mock_att = self._make_dummy_com_attachment(store_root)
            result = parse_attachment(store_root, mock_att, "test_record_key", "run1")
            assert result is not None
            assert result["extract_key"] != ""
            assert result["parse_status"] == "parsed"
            assert result["temp_deleted"] is True
            # Extract sidecar written
            extract = result["extract"]
            assert extract["_schema"] == "export.knowledgeExtract.v1"
            assert extract["parentRecordKey"] == "test_record_key"
            assert extract["audit"]["rawSourceRetained"] is False
            assert extract["audit"]["tempFileDeleted"] is True
            # Check content
            assert "Hello attachment!" in extract["content"]["text"]

    def test_unknown_binary_metadata_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store_root = os.path.join(tmp, "KnowledgeStore")
            os.makedirs(os.path.join(store_root, "temp", "parsing"), exist_ok=True)

            class MockBinAttachment:
                FileName = "data.bin"
                def SaveAsFile(self, path):
                    with open(path, "wb") as f:
                        f.write(b"\x00\x01\x02\x03")

            mock_att = MockBinAttachment()
            result = parse_attachment(store_root, mock_att, "rk1", "run1")
            assert result["parse_status"] == "metadata_only"
            assert result["needs_review"] is True
